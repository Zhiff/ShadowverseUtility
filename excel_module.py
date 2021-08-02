# -*- coding: utf-8 -*-
"""
Created on Fri Aug  7 21:42:37 2020
This in excel module. This module will handle almost all data manipulation in excel
@author: zhafi
"""
import pandas as pd
import openpyxl as oxl
import stat_helper as sh
from deckmodule import Deck


# This function will quickly convert all raw svportal links that found in excel document into deck archetype link. regardless of format
# Input : excel file
# Output : new excel file
def excel_convert_quick(excelfile, sheetname, custom=False):
    excel = oxl.load_workbook(excelfile)
    sheet = excel.worksheets[0]
    # Iterate all cells in excel sheet
    for column in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            #We only care about svportal, so it must be a string
            if type(cell.value) == str:
                if 'shadowverse-portal.com' in cell.value:
                    #determine archetype by calling checker function, save old value into pure hyperlink, save archetype as the string description
                    deck = Deck(cell.value)
                    archetype = deck.archetype_checker()
                    cell.hyperlink = cell.value
                    cell.value = archetype
    if (custom):
        excel.save(excelfile)
    else:
        excel.save('Excel_and_CSV/FilteredDecks_View.xlsx')

# This function will create FilteredDecks_Data xlsx which serves as main data processing. This file will be the basis for all statistics calculation
def excel_convert_dataset(svo_raw, maxdeck):
    df = pd.read_excel(svo_raw)
    for i in range(1, maxdeck+1):
        df[f'arc {i}'] = df[f'deck {i}'].apply(lambda x: Deck(x).archetype_checker())
    for i in range(1, maxdeck+1):
        df[f'class {i}'] = df[f'deck {i}'].apply(lambda x: Deck(x).class_checker_svo())
    if maxdeck == 3:
        df = sh.add_lineup_column_3decks(df)       
    elif maxdeck == 2:
        df = sh.add_lineup_column_2decks(df)
    elif maxdeck == 5:
        df = sh.add_lineup_column_5decks(df)
    
    writer = pd.ExcelWriter("Excel_and_CSV/FilteredDecks_Data.xlsx", options={'strings_to_urls': False})
    df.to_excel(writer, 'MainData')
    writer.save()

# This function is responsible to create Statistics and Breakdown file which is the one of the main output for this project.       
# requirements : name (lowercase) , deck 1 , deck 2 
def excel_statistics(filtered_data, maxdeck):
    
    #Adding a new column in df called lineup. Lineup is basically a list of 3 decks that has been sorted ex : {sword, dragon, blood}
    df = pd.read_excel(filtered_data)        
    outputfile = "Excel_and_CSV/Statistics and Breakdown.xlsx"
    # writer = pd.ExcelWriter(outputfile, options={'strings_to_urls': False})
    writer = pd.ExcelWriter(outputfile)
    if maxdeck == 3:
        df = sh.add_lineup_column_3decks(df)        
    elif maxdeck == 2:
        df = sh.add_lineup_column_2decks(df) 
    elif maxdeck == 5:
        df = sh.add_lineup_column_5decks(df)
    
    # Create a new dataframe that consists only lineup and their amount of occurrence
    lineup = sh.get_lineup_df(df, maxdeck)
    lineup.to_excel(writer, "Lineups")
    
    #Create a new dataframe that consists only deck and their amount of occurrence
    decks = sh.get_decks_df(df, maxdeck)
    decks.to_excel(writer, "Decks")
    
    # Breakdown each archetype
    tournament_breakdown(df, writer, maxdeck)  

    writer.save()
    
    # Freeze cardname and mean for deck breakdown
    statistics_freeze_highlight(outputfile)


#This function will handle Deck Archetype Breakdowns
#The sheet will list all players decklist grouped by archetype and compare it side by side in order to get bigger picture of the archetype
def tournament_breakdown(df, excelwriter, maxdeck):
    
    #Popular archetype filter. Will return the list of popular archetype with specified minimum occurrence
    occurrence = 1
    popular_archetype = sh.get_popular_archetype(df, occurrence, maxdeck)
    
    #Iterate each popular archetype
    for archetype in popular_archetype:
        flag_first = True   #Needed for first instance, resolve merge DF issue
        if archetype != 'Unknown Unknown': #Only proceed with valid data
            #Iterate the whole dataframe using i and x pointer
            for i in range(df.shape[0]): 
                for x in range(1,maxdeck+1):
                    # accessing arc x column data ( x = 1,2,3 )
                    decktype = df.loc[i,f'arc {x}']
                    if (decktype == archetype):
                        # accessing name column and svlink portal column, then extract the details using function
                        player_name = df.loc[i,'name'] 
                        decklist = df.loc[i,f'deck {x}']
                        details = Deck(decklist).deck_details()
                        if details is not None: #Validity Check
                            if flag_first == True:
                                #For the first instance, we simply initialize the data frame
                                arc_df = details.rename(columns={'Qty':player_name})
                                flag_first = False
                            else:
                                #Append dataframe with new dataframe
                                added_df = details.rename(columns={'Qty':player_name})
                                arc_df = pd.merge(arc_df, added_df, on='CardName', how='outer')
            
            # cleanup dataframe by filling NaN into 0
            arc_df = arc_df.fillna(0)
            arc_df = arc_df.set_index('CardName')
            
            #Add Mean, Median,and Standard Deviation into Dataframe
            arc_df = sh.add_statistics_tool(arc_df)
            
            #Reordering Columns, Mean column appears in front
            cols = list(arc_df.columns.values)
            arc_df = arc_df[[cols[-1]] + cols[0:-1]]
            arc_df = arc_df.sort_values('Average', ascending=False)
            arc_df.to_excel(excelwriter, archetype)

#This function will combine filtereddecks_view into stats and breakdown.xlsx for easier spreadsheets export
def combine_view_and_stats(viewfile, sheetname):
    file = 'Excel_and_CSV/Statistics and Breakdown.xlsx'
    excel1 = oxl.load_workbook(viewfile)
    excel2 = oxl.load_workbook(file)
    
    # assign source sheet and destination sheet
    src = excel1.worksheets[0]
    dst = excel2.create_sheet(title=sheetname, index=0)
    
    # copy all the decklist content in source to destination
    for row in src:
        for cell in row:
            dst[cell.coordinate].hyperlink = cell.hyperlink
            dst[cell.coordinate].value = cell.value
            
    excel2.save(file)

#This function will freeze first 2 column in statistics and highlight the important cards
def statistics_freeze_highlight(excelfile):
    excel = oxl.load_workbook(excelfile)
    breakdown = excel.sheetnames
    breakdown = breakdown[2:len(breakdown)]
    
    # Conditional Formatting, Highlight entire row if mean >= 2
    colorfill = oxl.styles.PatternFill(bgColor="A9A9A9")
    diffstyle = oxl.styles.differential.DifferentialStyle(fill=colorfill)
    rule = oxl.formatting.Rule(type='expression', dxf=diffstyle)
    rule.formula = ["$B2>=2"]

    for archetype in breakdown:
        sheet = excel[archetype]
        last_column = oxl.utils.cell.get_column_letter(sheet.max_column)
        sheet.freeze_panes = 'C1'
        sheet.conditional_formatting.add(f"A2:{last_column}80", rule)
        sheet.column_dimensions['A'].width = 30
    
    excel.save(excelfile)

#This function will handle conditional formatting for class coloring
def conditionalFormat(sheet):
    classmap = ['Forest', 'Sword', 'Rune', 'Dragon', 'Shadow', 'Blood', 'Haven', 'Portal']
    # classmap = ['엘프', '로얄', '위치', '드래곤', '네크로맨서', '뱀파이어', '비숍', '네메시스']
    colormap = ['E2EFDA', 'FFF2CC', 'CCCCFF', 'FCE4D6', 'FFCCFF', 'FFA39E', 'D0CECE', 'DDEBF7']
    # Repeat the conditional formatting assignment for each class
    for i in range(0,8):
    
        colorfill = oxl.styles.PatternFill(bgColor=colormap[i])
        diffstyle = oxl.styles.differential.DifferentialStyle(fill=colorfill)
        clmap = classmap[i]
        rule = oxl.formatting.Rule(type="containsText", operator="containsText", text=clmap, dxf=diffstyle)
        rule.formula = [f'NOT(ISERROR(SEARCH("{clmap}", A1)))']
        sheet.conditional_formatting.add("A1:F400", rule)

#This function will add color to View sheet and Decks sheet in Stats and Breakdown files
def add_class_color(mode):
# mode 1 : For Statistics and Breakdown sheet
# mode 2 : For Post SVO
    file = 'Excel_and_CSV/Statistics and Breakdown.xlsx'
    excel = oxl.load_workbook(file)
    sheet = excel.worksheets[0]
    conditionalFormat(sheet)
    if (mode == 1):
        sheet = excel.worksheets[1]
        conditionalFormat(sheet)
        sheet = excel['Decks']
        conditionalFormat(sheet)
        sheet = excel['Top16 Conversion']
        conditionalFormat(sheet)
        sheet = excel['Lineups']
        conditionalFormat(sheet)
    elif (mode == 2):
        sheet = excel.worksheets[1]
        conditionalFormat(sheet)
    elif (mode == 3):
        sheet = excel['Decks']
        conditionalFormat(sheet)
        sheet = excel['Lineups']
        conditionalFormat(sheet)
    
    
    excel.save(file)

def count_deck(filtered_data, maxdeck):
    df = pd.read_excel(filtered_data)
    decks = sh.get_decks_df(df, maxdeck)
    decks = decks[['Deck Archetype', 'Count']]
    
    return decks

def add_top16_names(top16df):
    file2 = 'Excel_and_CSV/FilteredDecks_Data.xlsx'
    file = 'Excel_and_CSV/JCGTop16_View.xlsx'
    
    group = pd.DataFrame({'Group':['Group 1','Group 2','Group 3','Group 4','Group 5','Group 6','Group 7','Group 8','Group 9','Group 10','Group 11','Group 12','Group 13','Group 14','Group 15','Group 16']})
    # top16df['arc 1'] = top16df['arc 1'].apply(lambda x: Deck(x).deckbuilder_converter())
    # top16df['arc 2'] = top16df['arc 2'].apply(lambda x: Deck(x).deckbuilder_converter())
    top16df = top16df.rename(columns={'arc 1':'deck 1', 'arc 2':'deck 2'})
    df = pd.concat([group, top16df],axis=1)
    
    
    writer = pd.ExcelWriter(file, options={'strings_to_urls': False})
    df.to_excel(writer, 'Qualified for Top16', index=False)
    writer.save()
    
    #Convert Top 16 decklinks back into Archetype and add it to statistics
    excel_convert_quick(file, 'Qualified for Top16', True)
    combine_view_and_stats(file, 'Qualified for Top16')
    

def add_conversion_rate(top16df):

    #Initialization
    file1 = 'Excel_and_CSV/Statistics and Breakdown.xlsx'
    book = oxl.load_workbook(file1)
   
    
    #count deck and combine with data
    top16df['arc 1'] = top16df['arc 1'].apply(lambda x: Deck(x).archetype_checker())
    top16df['arc 2'] = top16df['arc 2'].apply(lambda x: Deck(x).archetype_checker())
    decks = sh.deck_quick_count(top16df)
    decks['Top 16 Rep%'] = (round((decks['Count']/(int(top16df.shape[0])))*100, 2))
    decksdf = pd.read_excel(file1, sheet_name='Decks')
    decksdf = decksdf.rename(columns={'Count':'Total', 'Player %':'Group Rep%'})
    mergedeck = decks.merge(decksdf)
    mergedeck['Conversion Rate %'] = round(mergedeck['Count']/mergedeck['Total'], 4)*100 
    mergedeck = mergedeck.rename(columns={'Count':'Top 16', 'Total':'Group'})
    mergedeck = mergedeck.astype(str)
    mergedeck['Conversion Rate %'] = mergedeck['Conversion Rate %'].astype(float)
    mergedeck['Top 16 (Player%)'] = mergedeck['Top 16'] + ' (' + mergedeck['Top 16 Rep%'] + '%)'
    mergedeck['Group (Player%)'] = mergedeck['Group'] + ' (' + mergedeck['Group Rep%'] + '%)'
    mergedeck = mergedeck[['Deck Archetype','Top 16 (Player%)','Group (Player%)','Conversion Rate %']]
    
    writer = pd.ExcelWriter(file1, engine='openpyxl')
    writer.book = book
    mergedeck.to_excel(writer, sheet_name='Top16 Conversion')
    writer.save()
    
    currentorder=book.sheetnames
    myorder = neworder(currentorder, 4)
    book._sheets = [book._sheets[i] for i in myorder]
    book.save(file1)

#reordering sheets
def neworder(shlist, tpos):
    """Takes a list of ints, and inserts the last int, to tpos location (0-index)"""
    lst = []
    lpos = (len(shlist) - 1)
    # Just a counter
    for x in range(len(shlist)):
        if x > (tpos - 1) and x != tpos:
            lst.append(x-1)
        elif x == tpos:
            lst.append(lpos)
        else:
            lst.append(x)

    return lst

def convertSVOformat(svoexcel):
    carddb = "Excel_and_CSV/generatedURLcode.csv"
    # Create dictionary for class and hash->card name
    craftcode = { 'Forestcraft':'1' , 'Swordcraft':'2' , 'Runecraft':'3' , 'Dragoncraft':'4' , 'Shadowcraft':'5' , 'Bloodcraft':'6' , 'Havencraft':'7' , 'Portalcraft':'8' }
    dbdf = pd.read_csv(carddb)
    dbdf = dbdf.drop_duplicates(subset=['CardName'], keep='first')
    dbdf = dbdf.set_index('CardName')
    dbdf = dbdf['Code']
    carddict = dbdf.to_dict()
    # grab from SVO Excel. exact required fields : Player Name, Class, Card 1, Card 2, ... , Card 40 
    df = pd.read_excel(svoexcel)   
    df['Class'] = df['Class'].apply(lambda x: craftcode[x])
    # Change card name  -> Hash
    for i in range(2,42):
        df.iloc[:,i] = df.iloc[:,i].apply(lambda x: carddict[x])
    #Combine Hash
    df['rawdeck'] = ''
    for i in range(2,42):
        if i < 41:
            df['rawdeck'] = df['rawdeck'] + df.iloc[:,i] + '.'
        else:
            df['rawdeck'] = df['rawdeck'] + df.iloc[:,i]
    
    #Transform hash into hyperlink
    df['deck'] = 'https://shadowverse-portal.com/deck/3.' + df['Class'] + '.' + df['rawdeck'] + '?lang=en'
    df = df[['Player Name', 'deck']].rename(columns={'Player Name':'name'})
    
    #Combine decks that has same player name into one, then split them again so each player name has 3 decks column
    filtered = df.groupby(['name'])['deck'].apply(','.join).reset_index()
    filtered['deck 1'] = filtered['deck'].apply(lambda x: x.split(',')[0])
    filtered['deck 2'] = filtered['deck'].apply(lambda x: x.split(',')[1])
    filtered['deck 3'] = filtered['deck'].apply(lambda x: x.split(',')[2])
    
    #Previous method sorting name alphabetically, re-sort names according to SVO data
    final = filtered[['name','deck 1','deck 2','deck 3']]
    dfcopy = df['name'].drop_duplicates()
    sorteddf = pd.merge(dfcopy, final)
    
    # Save file into SVO.xlsc
    file = 'Excel_and_CSV/SVO.xlsx'
    writer = pd.ExcelWriter(file, options={'strings_to_urls': False})
    sorteddf.to_excel(writer, 'Qualified for Top16', index=False)
    writer.save()

